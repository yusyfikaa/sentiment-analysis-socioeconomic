import os
import time
import pandas as pd
from dotenv import load_dotenv
import google.generativeai as genai
from tqdm import tqdm
from openpyxl import load_workbook

# === Load API Key ===
load_dotenv()
api_key = os.getenv("GEMINI_API_KEY")
genai.configure(api_key=api_key)
model = genai.GenerativeModel(model_name="gemini-1.5-flash")

# === File paths ===
input_path = 'data/processed-data.xlsx'
output_path = 'data/topic_label.xlsx'

# Ensure output directory exists
os.makedirs(os.path.dirname(output_path), exist_ok=True)

# === Load dataset (no preprocessing) ===
df = pd.read_excel(input_path)

# === Create empty output file if it doesn't exist ===
if not os.path.exists(output_path):
    pd.DataFrame(columns=df.columns.tolist() + ['predicted_topic']).to_excel(output_path, index=False)
    print("🚀 Created empty output file to store results.")
else:
    print("🔁 Output file exists, resuming...")

# Determine how many rows are already processed
labeled = pd.read_excel(output_path)
start_index = len(labeled)
print(f"🔁 Resuming from row {start_index}")

# === Gemini labeling function ===
def predict_topic_from_comment(text, retries=3):
    prompt = f"""
You are an expert in socioeconomic impact analysis. You will analyze the 
following comment and classify it into one of the following three topics:

1. Employment Trend
2. Public Health
3. Cost of Living

Your task is to assign exactly one topic that best represents the main idea 
of the comment. Do not assign more than one topic. Return only the topic 
name exactly as it is (e.g., "Public Health"), without extra explanations.

Comment:
\"\"\"{text}\"\"\" 

Topic:
"""
    for attempt in range(retries):
        try:
            response = model.generate_content(prompt)
            label = response.text.strip()
            time.sleep(4.5)  # Respect rate limit

            allowed_labels = [
                "Employment Trend",
                "Public Health",
                "Cost of Living",
            ]
            predicted = label if label in allowed_labels else "Others"
            return predicted
        except Exception as e:
            print(f"[Attempt {attempt+1}] Error: {e}")
            time.sleep(5)
    return "Others"

# === Batch processing ===
batch_size = 50
total_rows = len(df)

for start in range(start_index, total_rows, batch_size):
    end = min(start + batch_size, total_rows)
    batch = df.iloc[start:end].copy()
    print(f"\n🔄 Processing rows {start} to {end - 1}")

    tqdm.pandas()
    batch['predicted_topic'] = batch['comment'].progress_apply(predict_topic_from_comment)

    # Append batch to output Excel directly
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        workbook = load_workbook(output_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1

        # Append without header
        batch.to_excel(writer, index=False, header=False, startrow=next_row - 1)

    print(f"✅ Saved rows {start} to {end - 1} to output.")
    time.sleep(10)  # Optional buffer between batches

print("All rows processed and saved successfully!")
